import openpyxl

def get_row_count(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    row_count = sheet.max_row
    return row_count

def get_column_count(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_column


def read_data(file,sheetname,rowno,colno):
    workbook = openpyxl.load_workbook(file)
    sheet =  workbook[sheetname]
    return sheet.cell(row=rowno,column=colno).value

def write_data(file,sheetname,rowno,colno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rowno,column=colno).value = data
    workbook.save(file)
